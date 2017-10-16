#! python3
# updateProduce.py - Corrects costs in produce sales spreadsheet.

import openpyxl
wb = openpyxl.load_workbook('E:\\MyCodes\\python\\python-auto\\tmp_file\\excel_files\\produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce types and their updated prices
PRICE_UPDATES = {
    'Carlic': 3.07,
    'Celery': 1.19,
    'Lemon': 1.27
}
# Loop through the rows and update the prices.
for rowNum in range(2, sheet.max_row):
    produceName = sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=2).value = PRICE_UPDATES[produceName]

wb.save('E:\\MyCodes\\python\\python-auto\\tmp_file\\excel_files\\update_files\\updateProduceSales.xlsx')











