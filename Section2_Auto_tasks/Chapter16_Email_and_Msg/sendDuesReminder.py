#! python3
# sendDuesReminders.py - Sends emails based on payment status in spreadsheet.

import openpyxl, smtplib, sys

# Open the spreadsheet and get the latest dues status.
wb = openpyxl.load_workbook('E:\\MyCodes\\python\\python-auto\\Section2_Auto_tasks\Chapter16_Email_and_Msg\\duesRecords.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.max_column
latestMonth = sheet.cell(row=1, column=lastCol)

# Check each member's payment status.
unpaidMembers = {}
for r in range(2, sheet.max_row + 1 ):
    payment = sheet.cell(row=r, column=lastCol).value
    if payment != 'paid':
        name = sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaidMembers[name] = email
# Log in to email account.
# smtpObj = smtplib.SMTP('smtp.sina.com', 587)
smtpObj = smtplib.SMTP_SSL('smtp.sina.com',465)
smtpObj.ehlo()
# smtpObj.starttls()
## ATTENTION!!##
# smtpObj.login('gaojz017@sina.com', sys.argv[1])
smtpObj.login('gaojz017@sina.com', sys.argv[1])
# Send out reminder emails.
for name, email in unpaidMembers.items():
    # body = "Subject: %s Python-test, dues unpaid. \nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you! " % (lastMonth, name, lastMonth)
    body = "Subject: %s dues unpaid.\nDear %s,\nRecords show that you have not paid dues for % s.Please make this payment as soon as possible.Thank you!" % (latestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus = smtpObj.sendmail('gaojz017@sina.com', email, body)

    if sendmailStatus != {}:
        print('There was a problem sending email to %s: %s' % (email, sendmailStatus))

smtpObj.quit()





